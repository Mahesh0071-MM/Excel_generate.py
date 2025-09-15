from pathlib import Path
import tempfile
import shutil
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import datetime



# Try to import matplotlib; if unavailable, charts will be skipped
try:
    import matplotlib.pyplot as plt
    HAVE_MPL = True
except Exception:
    HAVE_MPL = False
    # Do not exit; continue without charting

def auto_adjust_column_width(ws, min_width=8, max_width=50):
    """Auto-fit column widths for an openpyxl worksheet."""
    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
        except Exception:
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        width = max(min_width, min(max_length + 2, max_width))
        ws.column_dimensions[col_letter].width = width

def style_header_row(ws, header_fill_hex="4F81BD"):
    """Apply header style to first row of worksheet."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_fill_hex, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def generate_excel_report(csv_path=None,
                          excel_path=None,
                          use_file_dialog=False,
                          embed_charts=True):
    
    # Optionally import tkinter only when needed (to avoid issues in headless envs)
    if use_file_dialog:
        try:
            from tkinter import Tk, filedialog
        except Exception as e:
            raise RuntimeError("Tkinter unavailable. Run without GUI or install tkinter.") from e

    # Resolve csv_path
    if use_file_dialog and (csv_path is None):
        root = Tk()
        root.withdraw()
        p = filedialog.askopenfilename(title="Select CSV File", filetypes=[("CSV files", "*.csv"), ("All files","*.*")])
        root.destroy()
        if not p:
            raise FileNotFoundError("No CSV selected.")
        csv_path = p

    if csv_path is None:
        raise ValueError("csv_path must be provided (or set use_file_dialog=True).")

    csv_path = Path(csv_path).expanduser()
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    # Resolve excel_path
    if use_file_dialog and (excel_path is None):
        from tkinter import Tk, filedialog
        root = Tk()
        root.withdraw()
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")], title="Save Excel Report As")
        root.destroy()
        if not p:
            raise FileNotFoundError("No save location selected.")
        excel_path = p

    if excel_path is None:
        # Default to same folder as CSV
        excel_path = csv_path.with_name("Final_Report.xlsx")
    excel_path = Path(excel_path).expanduser()

    # Create a temporary directory for charts
    tmpdir = Path(tempfile.mkdtemp(prefix="excel_report_tmp_"))

    try:
        # Load CSV (robust reading)
        df = pd.read_csv(csv_path, low_memory=False)

        # Basic cleaning: drop unnamed columns created by Excel exports
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', case=False)]

        # If Year column is numeric with .0 values convert to ints when safe
        if "Year" in df.columns:
            try:
                # drop NA rows for conversion check
                if pd.api.types.is_float_dtype(df["Year"]) and df["Year"].dropna().apply(float.is_integer).all():
                    df["Year"] = df["Year"].astype("Int64")
                elif pd.api.types.is_object_dtype(df["Year"]):
                    # attempt converting strings like '2023.0'
                    df["Year"] = pd.to_numeric(df["Year"], errors="ignore")
            except Exception:
                pass

        # Group by Year (if present) and compute pivot-like summaries
        if "Year" in df.columns:
            grouped = df.groupby("Year")
            pivot_wind = grouped["Wind"].agg(Mean="mean", Max="max", Min="min").reset_index()
            pivot_temp = grouped["Temperature"].agg(Mean="mean", Max="max", Min="min").reset_index()
            if "Precipitation_mm" in df.columns:
                pivot_precip = grouped["Precipitation_mm"].agg(Total="sum").reset_index()
            else:
                pivot_precip = pd.DataFrame(columns=["Year", "Total"])
        else:
            pivot_wind = pd.DataFrame({"Message": ["No 'Year' column available"]})
            pivot_temp = pd.DataFrame({"Message": ["No 'Year' column available"]})
            pivot_precip = pd.DataFrame({"Message": ["No 'Year' column available"]})

        # Label counts
        label_counts = None
        if "Label" in df.columns:
            label_counts = df["Label"].value_counts().rename_axis("Label").reset_index(name="Count")

        # Missing value summary
        missing_summary = df.isna().sum().to_frame(name="MissingCount")
        missing_summary.index.name = "Column"
        missing_summary = missing_summary.reset_index()

        # Descriptive statistics
        stats = df.describe(include="all").transpose()

        # Write to Excel
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Raw Data")
            pivot_wind.to_excel(writer, index=False, sheet_name="Wind Summary")
            pivot_temp.to_excel(writer, index=False, sheet_name="Temperature Summary")
            pivot_precip.to_excel(writer, index=False, sheet_name="Precipitation Summary")
            stats.to_excel(writer, sheet_name="Statistics")
            missing_summary.to_excel(writer, index=False, sheet_name="Missing Values")
            if label_counts is not None:
                label_counts.to_excel(writer, index=False, sheet_name="Label Summary")

        # Load workbook for styling and embedding charts
        wb = load_workbook(excel_path)

        # Style all sheets and adjust widths
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                style_header_row(ws)
            except Exception:
                pass
            try:
                auto_adjust_column_width(ws)
            except Exception:
                pass

        # Create charts if requested and matplotlib is available
        chart_files = []
        if embed_charts and HAVE_MPL and "Year" in df.columns:
            try:
                # Wind trend (mean wind by year)
                wind_mean = df.groupby("Year")["Wind"].mean().reset_index()
                if not wind_mean.empty:
                    p = tmpdir / "wind_trend.png"
                    plt.figure(figsize=(8,4))
                    plt.plot(wind_mean["Year"], wind_mean["Wind"], marker="o")
                    plt.title("Average Wind Speed by Year")
                    plt.xlabel("Year")
                    plt.ylabel("Wind")
                    plt.grid(True, linestyle="--", alpha=0.4)
                    plt.tight_layout()
                    plt.savefig(p, dpi=150, bbox_inches="tight")
                    plt.close()
                    chart_files.append(p)

                # Precipitation total by year (if column exists)
                if "Precipitation_mm" in df.columns:
                    precip_total = df.groupby("Year")["Precipitation_mm"].sum().reset_index()
                    if not precip_total.empty:
                        p2 = tmpdir / "precip_total.png"
                        plt.figure(figsize=(8,4))
                        plt.bar(precip_total["Year"].astype(str), precip_total["Precipitation_mm"])
                        plt.title("Total Precipitation by Year")
                        plt.xlabel("Year")
                        plt.ylabel("Precipitation (mm)")
                        plt.tight_layout()
                        plt.savefig(p2, dpi=150, bbox_inches="tight")
                        plt.close()
                        chart_files.append(p2)

                # Add a Charts sheet then embed images there
                if chart_files:
                    if "Charts" in wb.sheetnames:
                        chart_ws = wb["Charts"]
                    else:
                        chart_ws = wb.create_sheet("Charts")
                    # Place images spaced downwards
                    row_offset = 2
                    for i, img_path in enumerate(chart_files):
                        img = ExcelImage(str(img_path))
                        anchor_cell = f"A{row_offset + i*25}"
                        chart_ws.add_image(img, anchor_cell)
            except Exception as e:
                # if chart creation or embedding fails, continue without crashing
                print("Warning: failed to create/insert charts:", e)

        # Save workbook after styling/embedding
        wb.save(excel_path)

        print(f"Excel report generated: {excel_path}")

        # Return path for convenience
        return str(excel_path)

    finally:
        # Cleanup tmpdir
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass


if __name__ == "__main__":
    csv_default = r"C:\Czekanowsky\Czekanowsky\czekanowsky\Datasets.csv"
    excel_default = r"C:\Users\Mahesh mm\Documents\Final_Report.xlsx"

    # Recommended: change to True if you prefer to pick files via GUI dialogs
    use_gui_dialogs = False

    # If matplotlib is not installed, charts will be skipped automatically.
    try:
        generate_excel_report(
            csv_path=csv_default if not use_gui_dialogs else None,
            excel_path=excel_default if not use_gui_dialogs else None,
            use_file_dialog=use_gui_dialogs,
            embed_charts=True
        )
    except Exception as err:
        print("ERROR:", err)
        sys.exit(1)

if __name__ == "__main__":
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"C:/Users/Mahesh mm/Documents/Final_Report_{timestamp}.xlsx"
    generate_excel_report(
        "C:/Czekanowsky/Czekanowsky/czekanowsky/Datasets.csv",
        output_path
    )
    print(f"✅ Report saved to: {output_path}")
